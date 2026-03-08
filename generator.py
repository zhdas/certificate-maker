import openpyxl
from openpyxl.drawing.image import Image
import win32com.client
import pythoncom
from PyPDF2 import PdfMerger
import os
import tempfile

def generate_full_package(uploaded_packing_path):
    pythoncom.CoInitialize()
    TEMPLATE_PATH = os.path.abspath("CERTIFICATE.xlsx")
    ADN_STAMP_PATH = os.path.abspath("ADN.png")
    ORIGINAL_STAMP_PATH = os.path.abspath("ORIGINAL.png")
    CM_TO_PX = 37.79

    wb_p = openpyxl.load_workbook(uploaded_packing_path, data_only=True)
    temp_dir = tempfile.mkdtemp()
    pdf_list = []
    excel_app = win32com.client.Dispatch("Excel.Application")
    excel_app.Visible = False

    mapping = {
        'B8': 'B8', 'B9': 'B9', 'B10': 'B10', 'B11': 'B11', 'B12': 'B12', 'B13': 'B13',
        'F3': 'F3', 'C17': 'B16', 'E8': 'E21', 'E15': 'E15', 'E16': 'E16',
        'B25': 'B24', 'G25': 'E24', 'B27': 'B26', 'B29': 'B28', 'B31': 'B30'
    }

    try:
        for sheet_name in wb_p.sheetnames:
            p_xlsx = os.path.join(temp_dir, f"p_{sheet_name}.xlsx")
            p_pdf = os.path.join(temp_dir, f"p_{sheet_name}.pdf")
            wb_p_temp = openpyxl.load_workbook(uploaded_packing_path)
            for sn in wb_p_temp.sheetnames:
                if sn != sheet_name: wb_p_temp.remove(wb_p_temp[sn])
            wb_p_temp.save(p_xlsx)
            doc_p = excel_app.Workbooks.Open(p_xlsx)
            doc_p.ExportAsFixedFormat(0, p_pdf)
            doc_p.Close(False)
            pdf_list.append(p_pdf)

            ws_p_data = wb_p[sheet_name]
            wb_c = openpyxl.load_workbook(TEMPLATE_PATH)
            ws_c = wb_c.active
            for src, dst in mapping.items():
                val = ws_p_data[src].value
                if src == 'G25' and val is not None:
                    ws_c[dst] = f"{val} KGS"
                else:
                    ws_c[dst] = val
            ws_c['F4'] = ws_p_data['F4'].value
            ws_c['B18'] = ws_p_data['F4'].value
            ws_c['E32'] = f"{ws_p_data['G25'].value} KGS" if ws_p_data['G25'].value else ""

            for img_p, cell, h, w in [(ADN_STAMP_PATH, 'E37', 2.8, 4.7), (ORIGINAL_STAMP_PATH, 'D25', 4.0, 9)]:
                img = Image(img_p)
                img.height, img.width = h * CM_TO_PX, w * CM_TO_PX
                ws_c.add_image(img, cell)

            c_xlsx = os.path.join(temp_dir, f"c_{sheet_name}.xlsx")
            c_pdf = os.path.join(temp_dir, f"c_{sheet_name}.pdf")
            wb_c.save(c_xlsx)
            doc_c = excel_app.Workbooks.Open(c_xlsx)
            doc_c.ExportAsFixedFormat(0, c_pdf)
            doc_c.Close(False)
            pdf_list.append(c_pdf)

        final_pdf = os.path.join(temp_dir, "FINAL_RESULT.pdf")
        merger = PdfMerger()
        for pdf in pdf_list: merger.append(pdf)
        merger.write(final_pdf)
        merger.close()
        return final_pdf
    finally:
        excel_app.Quit()
        pythoncom.CoUninitialize()