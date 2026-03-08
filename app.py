from pathlib import Path
import tempfile

import openpyxl
import streamlit as st

import design
from generator import generate_full_package


def save_uploaded_file(uploaded_file) -> str:
    suffix = Path(uploaded_file.name).suffix or '.xlsx'
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded_file.getbuffer())
        return tmp.name


def build_summary(xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    first_sheet = wb[wb.sheetnames[0]] if wb.sheetnames else None
    return {
        'sheets': len(wb.sheetnames),
        'company': first_sheet['B8'].value if first_sheet else '—',
    }


st.set_page_config(
    page_title='ADN SHORING — Генерация PDF',
    page_icon='📄',
    layout='centered',
)

design.set_page_style()
design.render_header()
design.render_instruction()

uploaded_file = st.file_uploader(
    'Загрузите PACKING LIST',
    type=['xlsx'],
    label_visibility='collapsed',
    help='Поддерживается формат .xlsx',
)

if uploaded_file:
    uploaded_path = save_uploaded_file(uploaded_file)
    summary = build_summary(uploaded_path)
    design.render_file_ready(uploaded_file.name, summary)

    generate = st.button('🚀 СГЕНЕРИРОВАТЬ PDF', use_container_width=True)

    if generate:
        with st.spinner('Система создаёт PDF из всех страниц...'):
            result_path = generate_full_package(uploaded_path)

        with open(result_path, 'rb') as pdf_file:
            st.success('PDF успешно создан. Теперь можно скачать готовый файл.')
            st.download_button(
                label='📥 СКАЧАТЬ PDF',
                data=pdf_file,
                file_name=f"{Path(uploaded_file.name).stem}.pdf",
                mime='application/pdf',
                use_container_width=True,
            )
else:
    design.render_upload_hint()
